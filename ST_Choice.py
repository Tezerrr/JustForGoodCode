import openpyxl
import random
import os, sys


def open_file(FN):
    with open(f"{FN}", mode="r", encoding="windows-1251") as f:
        FILE_NAME = f.read()
        return FILE_NAME


def write_file(FN):
    with open(f"Settings.txt", mode="w", encoding="windows-1251") as f:
        f.write(FN)


def set_file():
    files = os.listdir()
    for i in range(len(files)):
        print(f"[{i}] {files[i]}")
    while True:
        try:
            return files[int(input(f"Введите номер желаемого файла: "))]
        except:
            print("Попробуйте ещё раз!")


def set_sheet(wb):
    names = wb.sheetnames
    for i in range(len(names)):
        print(f"[{i}] {names[i]}")
    while True:
        try:
            return names[int(input("Выберите номер листа: "))]
        except:
            print("Попробуйте ещё раз!")


def set():
    FILE_NAME = set_file()
    print("Файл выбран - ", FILE_NAME)
    wb = openpyxl.load_workbook(filename=FILE_NAME)
    SHEET = set_sheet(wb)
    print("Лист выбран - ", SHEET)
    MODE = 0
    while MODE != "1" and MODE != "2":
        MODE = str(input("Имена и Фамилии в 1 или 2 колонках? (Введите 1 или 2) "))
    position = input(
        "Введите через пробел начальную букву и цифру, с которых идут имена (Например, 'B 3'): ").split()
    START_LETTER, START_NUMBER = str(position[0]), int(position[1])
    ALL = f"FILE_NAME: {FILE_NAME}\nSHEET: {SHEET}\nMODE: {MODE}\nSTART_LETTER: {START_LETTER}\nSTART_NUMBER: {START_NUMBER}"
    write_file(ALL)


def main():
    try:
        with open("Settings.txt", mode="r") as f:
            FILE_NAME = f.readline().split("FILE_NAME: ")[1][:-1]
            SHEET = f.readline().split("SHEET: ")[1][:-1]
            MODE = f.readline().split("MODE: ")[1][:-1]
            START_LETTER = f.readline().split("START_LETTER: ")[1][:-1]
            START_NUMBER = f.readline().split("START_NUMBER: ")[1]
    except:
        set()
        print("Настройка прошла успешно!")
    try:
        with open("Settings.txt", mode="r") as f:
            FILE_NAME = f.readline().split("FILE_NAME: ")[1][:-1]
            SHEET = f.readline().split("SHEET: ")[1][:-1]
            MODE = f.readline().split("MODE: ")[1][:-1]
            START_LETTER = f.readline().split("START_LETTER: ")[1][:-1]
            START_NUMBER = int(f.readline().split("START_NUMBER: ")[1])

        wb = openpyxl.load_workbook(filename=FILE_NAME)
        print(f"Выбран файл - {FILE_NAME}")
        print(f"Выбран лист - {SHEET}")
        print("Чтобы изменить лист - введите 1")
        print("Чтобы изменить настройки - введите 2")
        print("Чтобы приступить к опросу - просто Enter")
        ans = input()
        try:
            if int(ans) == 1:
                SHEET = set_sheet(wb)
                ALL = f"FILE_NAME:{FILE_NAME}\nSHEET:{SHEET}\nMODE: {MODE}\nSTART_LETTER: {START_LETTER}\nSTART_NUMBER: {START_NUMBER}"
                write_file(ALL)
            elif int(ans) == 2:
                set()
        except:
            pass
        names = []
        SHEET = wb[SHEET]
        while SHEET[f"{START_LETTER}{START_NUMBER}"].value:
            if MODE == "1":
                name = SHEET[f"{START_LETTER}{START_NUMBER}"].value
            elif MODE == "2":
                name = SHEET[f"{START_LETTER}{START_NUMBER}"].value + " " + SHEET[
                    f"{chr(ord(START_LETTER) + 1)}{START_NUMBER}"].value
            START_NUMBER += 1
            names.append(name)
        while len(names) > 0:
            name = names.pop(random.randrange(0, len(names)))
            print(f"{name} Осталось опросить: {len(names)}")
            print("Введите 0, если человек не ответил. В противном случае нажмите Enter.")
            try:
                command = int(input())
                if command == 0:
                    names.append(name)
                else:
                    continue
            except:
                continue

        print("Поздравляю, все сдали!")
        input()
    except Exception as ex:
        print(FILE_NAME)
        print("Ошибка при чтении файла.")
        print(ex)
        input()


if __name__ == "__main__":
    main()
